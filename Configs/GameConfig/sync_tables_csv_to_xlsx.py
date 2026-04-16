#!/usr/bin/env python3

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少依赖: openpyxl", file=sys.stderr)
    print("可先执行: python3 -m venv .venv_xlsx && . .venv_xlsx/bin/activate && pip install openpyxl", file=sys.stderr)
    sys.exit(1)


ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "Datas" / "__tables__.csv"
XLSX_PATH = ROOT / "Datas" / "__tables__.xlsx"


HEADER_ROW_1 = [
    "##var",
    "full_name",
    "value_type",
    "read_schema_from_file",
    "input",
    "index",
    "mode",
    "group",
    "comment",
    "output",
    "tags",
]

HEADER_ROW_2 = [
    "##",
    "全名(包含模块和名字)",
    "记录类名",
    "从excel读取定义",
    "文件列表",
    "表id字段",
    "模式",
    "分组",
    "注释",
    "输出文件",
    "",
]

HEADER_ROW_3 = [
    "##",
    "",
    "",
    "false时取已有定义，true为从excel标题头和属性栏读取定义",
    "可以多个，以逗号','分隔",
    '为空的话自动取value_type中第一个字段,多主键联合索引为key1+key2,多主键独立索引为"key1,key2"',
    "取值one|map|list，为空自动为map",
    "取值c|s|e，可以有多个，以逗号','分隔。空则表示属于所有分组",
    "",
    "",
    "",
]


def parse_bool(value: str):
    text = (value or "").strip().lower()
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no", ""}:
        return False if text else ""
    return value


def read_csv_rows():
    rows = []
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    if len(rows) < 5:
        raise ValueError("__tables__.csv 内容不足，至少需要 4 行表头和 1 行数据")
    return rows


def csv_row_to_xlsx_row(row):
    row = list(row) + [""] * (11 - len(row))
    return [
        row[0],                 # var
        row[1],                 # full_name
        row[2],                 # value_type
        parse_bool(row[7]),     # read_schema_from_file
        row[8],                 # input
        row[3],                 # index
        row[4],                 # mode
        row[5],                 # group
        row[6],                 # comment
        row[9],                 # output
        row[10],                # tags
    ]


def clear_sheet(ws):
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def write_row(ws, row_idx, values):
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx).value = value


def main():
    csv_rows = read_csv_rows()
    data_rows = csv_rows[4:]

    wb = load_workbook(XLSX_PATH)
    ws = wb[wb.sheetnames[0]]

    clear_sheet(ws)
    write_row(ws, 1, HEADER_ROW_1)
    write_row(ws, 2, HEADER_ROW_2)
    write_row(ws, 3, HEADER_ROW_3)

    written = 0
    for row_idx, row in enumerate(data_rows, start=4):
        if not any(cell.strip() for cell in row if isinstance(cell, str)):
            continue
        write_row(ws, row_idx, csv_row_to_xlsx_row(row))
        written += 1

    wb.save(XLSX_PATH)
    print(f"已同步 {written} 条记录: {CSV_PATH.name} -> {XLSX_PATH.name}")


if __name__ == "__main__":
    main()
